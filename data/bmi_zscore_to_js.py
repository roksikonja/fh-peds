"""
bmi_zscore_to_js.py

Converts BMI-SDS-LMS.xlsx into a compact JavaScript lookup table:
  website/bmi_zscore_table.js

The Excel file has two sheets:
  'Male=1'   — gender = 1
  'Female=2' — gender = 0 (form value) / 2 (sheet name)

Grid layout in each sheet:
  Rows: BMI values  10.00 → 50.00, step 0.05  (801 rows)
  Cols: age values  0.00  → 18.00, step 0.05  (361 cols)
  Cell: Z-score (SDS) for that (BMI, age) pair

Output format (bmi_zscore_table.js):
  Two flat Int16Array constants, one per sex, values = round(Z × 1000).
  Storing as integers keeps file size small (~1 MB) while retaining
  3 decimal-place precision (max Z error = 0.0005).

  Access pattern:
    bmi_row = clamp(round((bmi - 10) / 0.05), 0, 800)
    age_col = clamp(round(age        / 0.05), 0, 360)
    z = BMI_ZSCORE_MALE[bmi_row * 361 + age_col] / 1000

Usage:
  cd /home/rok/workspace/fh-peds/data
  ../ml-fh-peds/venv/bin/python bmi_zscore_to_js.py
"""

from __future__ import annotations

import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not found — run from the project venv:\n"
             "  ../ml-fh-peds/venv/bin/python bmi_zscore_to_js.py")

# ── Paths ────────────────────────────────────────────────────────────────────

HERE     = Path(__file__).parent
XLSX     = HERE / "BMI-SDS-LMS.xlsx"
OUT_JS   = HERE.parent / "website" / "bmi_zscore_table.js"

# ── Expected grid dimensions ─────────────────────────────────────────────────

BMI_MIN, BMI_MAX, BMI_STEP = 10.0, 50.0, 0.05
AGE_MIN, AGE_MAX, AGE_STEP =  0.0, 18.0, 0.05
N_BMI = round((BMI_MAX - BMI_MIN) / BMI_STEP) + 1   # 801
N_AGE = round((AGE_MAX - AGE_MIN) / AGE_STEP) + 1   # 361

SHEETS = {
    "Male=1":   "MALE",
    "Female=2": "FEMALE",
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_sheet(ws) -> list[int]:
    """Return a flat list of Z × 1000 integers, row-major (BMI outer, age inner)."""
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    age_cols  = [float(v) for v in header[1:]]
    bmi_rows  = [float(r[0]) for r in rows[1:]]

    if len(bmi_rows) != N_BMI:
        raise ValueError(f"Expected {N_BMI} BMI rows, got {len(bmi_rows)}")
    if len(age_cols) != N_AGE:
        raise ValueError(f"Expected {N_AGE} age cols, got {len(age_cols)}")

    # Verify step sizes
    actual_bmi_step = round(bmi_rows[1] - bmi_rows[0], 6)
    actual_age_step = round(age_cols[1]  - age_cols[0],  6)
    if abs(actual_bmi_step - BMI_STEP) > 1e-6:
        raise ValueError(f"Unexpected BMI step: {actual_bmi_step}")
    if abs(actual_age_step - AGE_STEP) > 1e-6:
        raise ValueError(f"Unexpected age step: {actual_age_step}")

    flat: list[int] = []
    nones = 0
    for row in rows[1:]:
        for v in row[1:]:
            if v is None:
                flat.append(0)
                nones += 1
            else:
                flat.append(round(float(v) * 1000))

    if nones:
        print(f"  Warning: {nones} None value(s) replaced with 0", file=sys.stderr)

    return flat


def clamp(v: int, lo: int, hi: int) -> int:
    return max(lo, min(hi, v))


def lookup(flat: list[int], bmi: float, age: float) -> float:
    """Bilinear-interpolated Z-score lookup from the flat array."""
    # Nearest-grid indices (no interpolation needed for step=0.05 data,
    # but we do linear interpolation for arbitrary inputs)
    bmi_f = (bmi - BMI_MIN) / BMI_STEP
    age_f = (age - AGE_MIN) / AGE_STEP

    b0 = clamp(int(bmi_f), 0, N_BMI - 2)
    a0 = clamp(int(age_f), 0, N_AGE - 2)
    b1 = b0 + 1
    a1 = a0 + 1

    tb = bmi_f - b0   # fractional position between b0 and b1
    ta = age_f - a0

    z00 = flat[b0 * N_AGE + a0] / 1000
    z01 = flat[b0 * N_AGE + a1] / 1000
    z10 = flat[b1 * N_AGE + a0] / 1000
    z11 = flat[b1 * N_AGE + a1] / 1000

    return (z00 * (1 - tb) * (1 - ta) +
            z01 * (1 - tb) * ta +
            z10 * tb * (1 - ta) +
            z11 * tb * ta)


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    if not XLSX.exists():
        sys.exit(f"Input file not found: {XLSX}")

    print(f"Reading {XLSX} …")
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

    arrays: dict[str, list[int]] = {}
    for sheet_name, const_name in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            sys.exit(f"Sheet {sheet_name!r} not found in workbook")
        print(f"  Loading sheet {sheet_name!r} → BMI_ZSCORE_{const_name} …")
        arrays[const_name] = load_sheet(wb[sheet_name])
        print(f"    {len(arrays[const_name])} values "
              f"({N_BMI} BMI × {N_AGE} age)")

    # ── Sanity checks ────────────────────────────────────────────────────────
    print("Running sanity checks …")

    # Male, age=7.3, BMI=17.5 → expected ~1.12  (verified interactively)
    z = lookup(arrays["MALE"], bmi=17.5, age=7.3)
    assert abs(z - 1.12) < 0.05, f"Male age=7.3 BMI=17.5: expected ~1.12, got {z:.3f}"
    print(f"  Male   age=7.3  BMI=17.5 → Z={z:.3f}  ✓")

    # Male, age=7.3, BMI≈15.6 → expected ~0  (median)
    z0 = lookup(arrays["MALE"], bmi=15.6, age=7.3)
    assert abs(z0) < 0.05, f"Male age=7.3 BMI=15.6: expected ~0, got {z0:.3f}"
    print(f"  Male   age=7.3  BMI=15.6 → Z={z0:.3f}  ✓")

    # Female, age=10.0, BMI=10.0 → should be a large negative Z
    zf = lookup(arrays["FEMALE"], bmi=10.0, age=10.0)
    assert zf < -2.5, f"Female age=10 BMI=10: expected < -2.5, got {zf:.3f}"
    print(f"  Female age=10.0 BMI=10.0 → Z={zf:.3f}  ✓")

    # Boundary: clamp to grid
    z_edge = lookup(arrays["MALE"], bmi=50.0, age=18.0)
    print(f"  Male   age=18.0 BMI=50.0 → Z={z_edge:.3f}  (boundary)")

    # ── Write JS ─────────────────────────────────────────────────────────────
    print(f"\nWriting {OUT_JS} …")
    OUT_JS.parent.mkdir(parents=True, exist_ok=True)

    lines: list[str] = [
        "// AUTO-GENERATED by data/bmi_zscore_to_js.py — do not edit by hand.",
        "// Source: data/BMI-SDS-LMS.xlsx (UK90 reference data)",
        "//",
        "// BMI Z-score (SDS) lookup table for children aged 0–18.",
        "// Two flat Int16Array constants, one per sex:",
        "//   BMI_ZSCORE_MALE   — gender = 1 (male)",
        "//   BMI_ZSCORE_FEMALE — gender = 0 (female)",
        "//",
        f"// Grid: BMI {BMI_MIN}–{BMI_MAX} step {BMI_STEP} ({N_BMI} rows)",
        f"//       age {AGE_MIN}–{AGE_MAX} step {AGE_STEP} ({N_AGE} cols), row-major",
        "// Values = round(Z × 1000)  →  divide by 1000 to recover Z.",
        "//",
        "// Lookup (nearest grid point):",
        "//   bmi_row = Math.min(Math.max(Math.round((bmi - 10) / 0.05), 0), 800)",
        "//   age_col = Math.min(Math.max(Math.round( age        / 0.05), 0), 360)",
        "//   z = table[bmi_row * 361 + age_col] / 1000",
        "//",
        "// Use bmiToZScore(bmi, age, gender) for bilinear interpolation.",
    ]

    for const_name, flat in arrays.items():
        csv = ",".join(map(str, flat))
        lines.append(f"\nconst BMI_ZSCORE_{const_name} = new Int16Array([{csv}]);")

    lines.append(r"""
/**
 * Convert raw BMI to age- and sex-adjusted Z-score (SDS) using the UK90
 * reference data (BMI-SDS-LMS.xlsx).
 *
 * Uses bilinear interpolation within the lookup grid.
 * Inputs outside the grid (BMI < 10, BMI > 50, age > 18) are clamped.
 *
 * @param {number} bmi     Body Mass Index in kg/m²
 * @param {number} age     Age in years (0–18)
 * @param {number} gender  0 = Female, 1 = Male  (matches form/model convention)
 * @returns {number}       BMI Z-score (SDS)
 */
function bmiToZScore(bmi, age, gender) {
  const table = gender === 1 ? BMI_ZSCORE_MALE : BMI_ZSCORE_FEMALE;

  const N_AGE = 361;
  const BMI_MIN = 10, BMI_STEP = 0.05, N_BMI = 801;
  const AGE_MIN =  0, AGE_STEP = 0.05, N_AGE_MAX = 360;

  const bmiF = (bmi - BMI_MIN) / BMI_STEP;
  const ageF = (age - AGE_MIN) / AGE_STEP;

  const b0 = Math.min(Math.max(Math.floor(bmiF), 0), N_BMI - 2);
  const a0 = Math.min(Math.max(Math.floor(ageF), 0), N_AGE_MAX - 1);
  const b1 = b0 + 1;
  const a1 = Math.min(a0 + 1, N_AGE_MAX);

  const tb = bmiF - b0;
  const ta = ageF - a0;

  const z00 = table[b0 * N_AGE + a0] / 1000;
  const z01 = table[b0 * N_AGE + a1] / 1000;
  const z10 = table[b1 * N_AGE + a0] / 1000;
  const z11 = table[b1 * N_AGE + a1] / 1000;

  return z00 * (1 - tb) * (1 - ta)
       + z01 * (1 - tb) * ta
       + z10 * tb        * (1 - ta)
       + z11 * tb        * ta;
}""")

    OUT_JS.write_text("\n".join(lines) + "\n")
    size_kb = OUT_JS.stat().st_size / 1024
    print(f"  Done. {size_kb:.0f} KB written.")


if __name__ == "__main__":
    main()
